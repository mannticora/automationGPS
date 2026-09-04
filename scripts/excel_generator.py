"""Generación del reporte Excel de validaciones GPS."""
from datetime import datetime
from pathlib import Path

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import Config

STREET_VIEW_COLUMN = "Verificación Street View"

_STREET_VIEW_VERDICT_LABELS = {
    "coincide": "✓ Coincide",
    "coincide_parcial": "≈ Coincide parcial",
    "no_coincide": "✗ No coincide",
    "sin_street_view": "— Sin cobertura Street View",
    "indeterminado": "? Indeterminado",
}

HEADERS = [
    "Case ID",
    "Nombre Negocio",
    "Nombre Negocio (Original)",
    "GPS Actual (lat, lon)",
    "GPS Corregido (lat, lon)",
    "Estado",
    "Distancia Error (m)",
    "Enlace Google Maps",
    "Notas",
    "Estatus Negocio",
    "Marcas Registradas",
    "Fotos Faltantes",
    "Calidad Sugerida",
    "Tipo Encuesta Sugerido",
    "Observaciones",
]

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_FILL_VALIDADO = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
_FILL_CORRECCION = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_ERROR = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _fill_for_status(status: str) -> PatternFill:
    """Elige el color de fila según el estado de validación del caso."""
    if "Validado" in status:
        return _FILL_VALIDADO
    if "corrección" in status:
        return _FILL_CORRECCION
    return _FILL_ERROR


def _format_missing_photos(missing_photos: list | None, total_photo_fields: int) -> str:
    """Formatea la lista de fotos faltantes para la columna del Excel."""
    if not total_photo_fields:
        return ""
    missing_photos = missing_photos or []
    if not missing_photos:
        return "Ninguna ✓"
    if len(missing_photos) == total_photo_fields:
        return f"Todas ({total_photo_fields}/{total_photo_fields})"
    return f"{', '.join(missing_photos)} ({len(missing_photos)}/{total_photo_fields})"


def generate_excel_report(cases: list[dict], output_filename: str | None = None) -> str:
    """Genera el Excel de validaciones a partir de la lista de casos procesados.

    `output_filename` puede ser un nombre simple (se guarda bajo `Config.OUTPUT_DIR`)
    o una ruta absoluta. Si no se provee, se usa `Config.EXCEL_FILENAME_PATTERN` con
    la fecha/hora actual. Devuelve la ruta final del archivo generado.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Validaciones"

    worksheet.append(HEADERS)
    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for case in cases:
        gps_actual = case.get("gps_actual")
        gps_corregido = case.get("gps_corregido")
        maps_link = case.get("maps_link", "")
        worksheet.append([
            case.get("case_id", ""),
            case.get("business_name", ""),
            case.get("business_name_original", ""),
            f"{gps_actual[0]}, {gps_actual[1]}" if gps_actual else "",
            f"{gps_corregido[0]}, {gps_corregido[1]}" if gps_corregido else "",
            case.get("status", ""),
            case.get("distance_error", ""),
            f'=HYPERLINK("{maps_link}", "Ver en Maps")' if maps_link else "",
            case.get("notes", ""),
            case.get("business_status", ""),
            ", ".join(case.get("brands", [])) or "",
            _format_missing_photos(case.get("missing_photos"), case.get("total_photo_fields", 0)),
            case.get("calidad_sugerida", ""),
            case.get("tipo_encuesta_sugerido", ""),
            case.get("observaciones_calidad", ""),
        ])

    status_column = HEADERS.index("Estado") + 1
    for row_idx in range(2, worksheet.max_row + 1):
        status_value = str(worksheet.cell(row=row_idx, column=status_column).value or "")
        fill = _fill_for_status(status_value)
        for cell in worksheet[row_idx]:
            cell.fill = fill

    for column_cells in worksheet.columns:
        max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 50)

    if not output_filename:
        output_filename = datetime.now().strftime(Config.EXCEL_FILENAME_PATTERN)

    output_path = Path(output_filename)
    if not output_path.is_absolute() and output_path.parent == Path("."):
        output_path = Path(Config.OUTPUT_DIR) / output_filename
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook.save(output_path)
    logger.success(f"Excel guardado: {output_path}")
    return str(output_path)


def append_street_view_verdicts(excel_path: str, verdicts: list[dict]) -> None:
    """Agrega (o actualiza) la columna "Verificación Street View" en un Excel ya generado.

    `verdicts` es la lista de resultados del workflow `verificar-street-view` (uno por
    caso): `{"case_id": str, "street_view_available": bool, "visible_signage": str,
    "match_verdict": str, "notes": str}`. Empareja por Case ID (columna "Case ID") y dejar
    en blanco las filas de casos no incluidos en `verdicts`. No vuelve a correr la
    verificación por sí sola: es un paso manual/asistido (ver docs/FLUJO_TRABAJO.md).
    """
    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    header_row = [cell.value for cell in worksheet[1]]
    if HEADERS.index("Case ID") + 1 != 1:
        raise ValueError("Se esperaba que 'Case ID' fuera la primera columna del Excel.")

    if STREET_VIEW_COLUMN in header_row:
        sv_column = header_row.index(STREET_VIEW_COLUMN) + 1
    else:
        sv_column = worksheet.max_column + 1
        header_cell = worksheet.cell(row=1, column=sv_column, value=STREET_VIEW_COLUMN)
        header_cell.fill = _HEADER_FILL
        header_cell.font = _HEADER_FONT
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    verdicts_by_case_id = {str(v["case_id"]): v for v in verdicts}
    for row_idx in range(2, worksheet.max_row + 1):
        case_id = str(worksheet.cell(row=row_idx, column=1).value or "")
        verdict = verdicts_by_case_id.get(case_id)
        if not verdict:
            continue
        label = _STREET_VIEW_VERDICT_LABELS.get(verdict["match_verdict"], verdict["match_verdict"])
        signage = verdict.get("visible_signage") or ""
        text = f"{label}" + (f" — rótulo visible: \"{signage}\"" if signage else "") + f". {verdict.get('notes', '')}"
        worksheet.cell(row=row_idx, column=sv_column, value=text.strip())

    worksheet.column_dimensions[worksheet.cell(row=1, column=sv_column).column_letter].width = 60
    workbook.save(excel_path)
    logger.success(f"Columna '{STREET_VIEW_COLUMN}' agregada/actualizada en: {excel_path}")
